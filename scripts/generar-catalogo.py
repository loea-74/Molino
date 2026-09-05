"""
Genera src/content/catalogo.json, que es lo que la página enseña.

    python scripts/generar-catalogo.py

Lee dos archivos, con papeles distintos a propósito:

  Plantilla_Productos.xlsx   el export del punto de venta. Dice qué EXISTE.
  Catalogo_Control.xlsx      opcional. Dice qué se ENSEÑA, con qué nombre y en
                             qué orden. Se crea con scripts/crear-control.py.

Si no hay archivo de control, sale todo con los nombres limpiados
automáticamente. Si lo hay, manda él: sus nombres pisan a los automáticos y lo
marcado como NO no aparece.

El export viene tal como se teclea en caja: TODO EN MAYUSCULAS, sin acentos,
con "OTROS" y "OTRAS" como categorías distintas y códigos internos sueltos. La
limpieza vive en generar_catalogo_comun.py, compartida con crear-control.py
para que las dos herramientas digan lo mismo.
"""

import json
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from generar_catalogo_comun import bonito, es_codigo, leer_export  # noqa: E402

RAIZ = Path(__file__).resolve().parent.parent
CONTROL = RAIZ / "Catalogo_Control.xlsx"
SALIDA = RAIZ / "src" / "content" / "catalogo.json"

#: Cuántos productos se enseñan de cada categoría cuando el control no dice otra cosa.
POR_CATEGORIA = 5

#: Los servicios de molienda, ya consolidados. El export los trae repetidos por
#: variantes de cobro ("Molienda 45", "Molienda Maiz Kilo"), que dentro de la
#: caja distinguen tarifas pero de cara al cliente son el mismo servicio.
SERVICIOS = [
    {"nombre": "Molienda de maíz", "nota": "Nixtamalizado, para pozole o para masa"},
    {"nombre": "Nixtamalización y molienda", "nota": "El proceso completo, de grano a masa"},
    {"nombre": "Molienda de café", "nota": "Al punto que pidas"},
    {"nombre": "Molienda de especias y granos", "nota": "Molemos lo que traigas"},
    {"nombre": "Molienda de moles", "nota": "Pastas de mole a tu receta"},
    {"nombre": "Molienda de pan", "nota": "Pan molido fresco"},
]


def leer_control():
    """Las decisiones del archivo de control, si existe."""
    if not CONTROL.exists():
        return {}, {}, {}

    libro = openpyxl.load_workbook(CONTROL, data_only=True)

    def hoja(nombre, claves):
        if nombre not in libro.sheetnames:
            return {}
        h = libro[nombre]
        cab = [c.value for c in h[1]]
        salida = {}
        for fila in h.iter_rows(min_row=2, values_only=True):
            if not fila or fila[0] is None:
                continue
            reg = dict(zip(cab, fila))
            try:
                salida[tuple(str(reg[k]) for k in claves)] = reg
            except KeyError:
                continue
        return salida

    return (
        hoja("Menu", ["Clave del departamento"]),
        hoja("Categorias", ["Clave del departamento", "Clave de la categoria"]),
        hoja("Productos", ["Clave"]),
    )


def lo_editado_a_mano():
    """
    Lo que se haya corregido desde el panel, indexado por clave.

    Sin esto, cada export nuevo borraría el trabajo hecho a mano: "PIÑON" se
    volvería a limpiar a "Piñon" y quien lo arregló tendría que arreglarlo otra
    vez. Se cruza por clave y no por nombre, precisamente porque el nombre es
    lo que cambia.
    """
    if not SALIDA.exists():
        return {}, {}, {}
    try:
        actual = json.loads(SALIDA.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return {}, {}, {}

    prods, cats, deps = {}, {}, {}
    for d in actual.get("departamentos", []):
        if "clave" not in d:
            continue  # formato viejo, sin claves: no hay nada que cruzar
        deps[d["clave"]] = {"nombre": d.get("nombre"), "oculto": d.get("oculto", False)}
        for c in d.get("categorias", []):
            cats[(d["clave"], c.get("clave"))] = {
                "nombre": c.get("nombre"), "oculto": c.get("oculto", False),
                "cuantos": c.get("cuantos"),
            }
            for x in c.get("productos", []):
                prods[x["clave"]] = {
                    "nombre": x.get("nombre"), "oculto": x.get("oculto", False),
                    "destacado": x.get("destacado", False),
                }
    return prods, cats, deps


def si(valor, por_omision=True):
    """Lee un SI/NO del archivo de control. Vacío = lo que diga por omisión."""
    if valor is None or str(valor).strip() == "":
        return por_omision
    return str(valor).strip().upper().startswith("S")


def texto(valor):
    return (str(valor).strip() if valor is not None else "")


def main() -> None:
    productos, _ = leer_export()
    if not productos:
        sys.exit("El export no trajo productos. ¿Está Plantilla_Productos.xlsx en su sitio?")

    ctrl_menu, ctrl_cats, ctrl_prods = leer_control()
    hay_control = bool(ctrl_menu or ctrl_cats or ctrl_prods)
    ed_prods, ed_cats, ed_deps = lo_editado_a_mano()
    conservados = 0
    ocultos = {"productos": 0, "categorias": 0, "departamentos": 0}
    renombrados = 0

    arbol = defaultdict(lambda: defaultdict(list))
    for p in productos:
        reg = ctrl_prods.get((p["clave"],), {})
        if not si(reg.get("Mostrar")):
            ocultos["productos"] += 1
            continue
        # Manda el Excel de control; si no dice nada, lo corregido desde el
        # panel; y si tampoco, el nombre limpiado automáticamente.
        editado = ed_prods.get(p["clave"], {})
        nombre = texto(reg.get("Nombre en la web")) or editado.get("nombre") or p["nombre"]
        if nombre != p["nombre"]:
            renombrados += 1
            if not texto(reg.get("Nombre en la web")):
                conservados += 1
        if editado.get("oculto") and si(reg.get("Mostrar")):
            ocultos["productos"] += 1
            continue
        arbol[p["departamento"]][p["categoria"]].append({
            "clave": p["clave"],
            "nombre": nombre,
            "destacado": si(reg.get("Destacar"), editado.get("destacado", p["favorito"])),
        })

    departamentos = []
    for dep in arbol:
        reg_dep = ctrl_menu.get((dep,), {})
        if ed_deps.get(dep, {}).get("oculto"):
            ocultos["departamentos"] += 1
            continue
        if not si(reg_dep.get("Mostrar")):
            ocultos["departamentos"] += 1
            continue

        categorias = []
        for cat, prods in arbol[dep].items():
            reg_cat = ctrl_cats.get((dep, cat), {})
            if ed_cats.get((dep, cat), {}).get("oculto"):
                ocultos["categorias"] += 1
                continue
            if not si(reg_cat.get("Mostrar"), not es_codigo(cat)):
                ocultos["categorias"] += 1
                continue
            try:
                cuantos = max(1, int(reg_cat.get("Cuantos productos enseñar") or POR_CATEGORIA))
            except (TypeError, ValueError):
                cuantos = POR_CATEGORIA
            # primero lo destacado, luego alfabético
            orden = sorted(prods, key=lambda x: (not x["destacado"], x["nombre"]))
            categorias.append({
                "clave": cat,
                "nombre": texto(reg_cat.get("Nombre en la web")) or bonito(cat),
                "cuantos": cuantos,
                "productos": orden,
                # "Otros" al final: es el cajón de sastre, no una categoría real
                "_orden": (cat == "OTROS", -len(prods), cat),
            })

        if not categorias:
            continue
        categorias.sort(key=lambda c: c["_orden"])
        total = sum(len(c["productos"]) for c in categorias)
        orden_dep = reg_dep.get("Orden")
        departamentos.append({
            "clave": dep,
            "nombre": texto(reg_dep.get("Nombre en la web")) or bonito(dep),
            "categorias": [
                OrderedDict([
                    ("clave", c["clave"]),
                    ("nombre", c["nombre"]),
                    ("cuantos", c["cuantos"]),
                    # van TODOS, no sólo los de muestra: así el panel puede
                    # buscar y corregir cualquiera, y esconder uno deja que
                    # el siguiente ocupe su sitio en vez de dejar el hueco
                    ("productos", [
                        OrderedDict([("clave", x["clave"]), ("nombre", x["nombre"])]
                                    + ([("destacado", True)] if x["destacado"] else []))
                        for x in c["productos"]
                    ]),
                ])
                for c in categorias
            ],
            "_orden": (orden_dep if isinstance(orden_dep, int) else 999, -total),
        })

    departamentos.sort(key=lambda d: d["_orden"])
    limpios = [
        OrderedDict([("clave", d["clave"]), ("nombre", d["nombre"]), ("categorias", d["categorias"])])
        for d in departamentos
    ]

    SALIDA.write_text(
        json.dumps(
            OrderedDict([
                ("generado", "scripts/generar-catalogo.py"),
                ("servicios", SERVICIOS),
                ("departamentos", limpios),
            ]),
            ensure_ascii=False, indent=2,
        ) + "\n",
        encoding="utf-8",
    )

    total = sum(len(c["productos"]) for d in limpios for c in d["categorias"])
    muestra = sum(min(c["cuantos"], len(c["productos"])) for d in limpios for c in d["categorias"])
    cats = sum(len(d["categorias"]) for d in limpios)
    print(f"{SALIDA.relative_to(RAIZ)}")
    print(f"  {len(limpios)} departamentos · {cats} categorías · {total} productos")
    print(f"  {muestra} de muestra · {len(SERVICIOS)} servicios")
    if conservados:
        print(f"  {conservados} nombres corregidos a mano se conservaron")
    if hay_control:
        print(f"  control aplicado: {renombrados} renombrados · ocultos "
              f"{ocultos['productos']} productos, {ocultos['categorias']} categorías, "
              f"{ocultos['departamentos']} departamentos")
    else:
        print("  (sin archivo de control: sale todo con los nombres automáticos)")


if __name__ == "__main__":
    main()
