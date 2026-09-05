"""
Crea Catalogo_Control.xlsx: el archivo con el que se decide qué sale en la web.

    python scripts/crear-control.py

Reparto de papeles, para que no haya dos verdades:

  Plantilla_Productos.xlsx   el export del punto de venta. Dice qué EXISTE.
                             No se toca a mano: se reemplaza por uno nuevo.
  Catalogo_Control.xlsx      este archivo. Dice qué se ENSEÑA, con qué nombre
                             y en qué orden. Este sí se edita.

Al volver a correrlo, conserva lo que ya estaba decidido y sólo añade los
renglones nuevos que hayan aparecido en el export. Nunca pisa una decisión.
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from generar_catalogo_comun import bonito, es_codigo, leer_export  # noqa: E402

CONTROL = RAIZ / "Catalogo_Control.xlsx"

CABECERA = PatternFill("solid", fgColor="3A2010")
EDITABLE = PatternFill("solid", fgColor="FDF6E8")
REFERENCIA = PatternFill("solid", fgColor="F0EDE8")
BLANCA = Font(color="FBF6ED", bold=True, size=11)
GRIS = Font(color="8A8580", size=10)


def hoja(libro, titulo, columnas, filas, editables_desde):
    """Una hoja con cabecera oscura, anchos sensatos y las columnas que se
    pueden editar marcadas en crema. Las de referencia van en gris: son el
    espejo del punto de venta y tocarlas no sirve de nada."""
    h = libro.create_sheet(titulo)
    h.append([c[0] for c in columnas])
    for i, (nombre, ancho) in enumerate(columnas, start=1):
        celda = h.cell(row=1, column=i)
        celda.fill = CABECERA
        celda.font = BLANCA
        celda.alignment = Alignment(vertical="center", wrap_text=True)
        h.column_dimensions[get_column_letter(i)].width = ancho
    h.row_dimensions[1].height = 30

    for fila in filas:
        h.append(fila)

    for r in range(2, h.max_row + 1):
        for c in range(1, len(columnas) + 1):
            celda = h.cell(row=r, column=c)
            if c < editables_desde:
                celda.fill = REFERENCIA
                celda.font = GRIS
            else:
                celda.fill = EDITABLE

    h.freeze_panes = "A2"
    return h


def si_no(h, columna, primera, ultima):
    """Lista desplegable SI/NO, para que nadie escriba 'si ', 'x' o 'Sí'."""
    v = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    v.error = "Escribe SI o NO"
    v.errorTitle = "Sólo SI o NO"
    h.add_data_validation(v)
    v.add(f"{columna}{primera}:{columna}{ultima}")


def previas():
    """Lo ya decidido en el archivo anterior, para no perderlo al regenerar."""
    if not CONTROL.exists():
        return {}, {}, {}
    libro = openpyxl.load_workbook(CONTROL)
    def leer(nombre, clave_cols):
        if nombre not in libro.sheetnames:
            return {}
        h = libro[nombre]
        cab = [c.value for c in h[1]]
        out = {}
        for fila in h.iter_rows(min_row=2, values_only=True):
            if not fila or fila[0] is None:
                continue
            clave = tuple(str(fila[cab.index(c)]) for c in clave_cols)
            out[clave] = dict(zip(cab, fila))
        return out
    return (
        leer("Menu", ["Clave del departamento"]),
        leer("Categorias", ["Clave del departamento", "Clave de la categoria"]),
        leer("Productos", ["Clave"]),
    )


def main():
    productos, servicios = leer_export()
    if not productos:
        sys.exit("El export no trajo productos.")

    menu_previo, cats_previas, prods_previos = previas()

    # ── el menú: los departamentos ──
    deps = {}
    for p in productos:
        deps.setdefault(p["departamento"], 0)
        deps[p["departamento"]] += 1

    filas_menu = []
    for i, (dep, n) in enumerate(sorted(deps.items(), key=lambda x: -x[1]), start=1):
        antes = menu_previo.get((dep,), {})
        filas_menu.append([
            dep, n,
            antes.get("Nombre en la web") or bonito(dep),
            antes.get("Mostrar") or "SI",
            antes.get("Orden") or i,
        ])

    # ── las categorías ──
    cats = {}
    for p in productos:
        clave = (p["departamento"], p["categoria"])
        cats[clave] = cats.get(clave, 0) + 1

    filas_cat = []
    for (dep, cat), n in sorted(cats.items(), key=lambda x: (x[0][0], -x[1])):
        antes = cats_previas.get((dep, cat), {})
        filas_cat.append([
            dep, cat, n,
            antes.get("Nombre en la web") or bonito(cat),
            antes.get("Mostrar") or ("NO" if es_codigo(cat) else "SI"),
            antes.get("Cuantos productos enseñar") or 5,
        ])

    # ── los productos ──
    filas_prod = []
    for p in sorted(productos, key=lambda x: (x["departamento"], x["categoria"], x["nombre"])):
        antes = prods_previos.get((p["clave"],), {})
        filas_prod.append([
            p["clave"], p["departamento"], p["categoria"], p["nombre_original"],
            antes.get("Nombre en la web") or p["nombre"],
            antes.get("Mostrar") or "SI",
            antes.get("Destacar") or ("SI" if p["favorito"] else "NO"),
        ])

    libro = openpyxl.Workbook()
    libro.remove(libro.active)

    # ── instrucciones, en la primera hoja ──
    guia = libro.create_sheet("Cómo se usa")
    guia.column_dimensions["A"].width = 105
    lineas = [
        ("Cómo se usa este archivo", True),
        ("", False),
        ("Aquí se decide QUÉ SALE en el catálogo de la página y CON QUÉ NOMBRE.", False),
        ("Lo que EXISTE sale del export del punto de venta; eso no se edita aquí.", False),
        ("", False),
        ("Las columnas grises son de referencia: reflejan el punto de venta.", False),
        ("Cambiarlas no sirve de nada — al regenerar se vuelven a leer de ahí.", False),
        ("Las columnas en crema son las que sí se editan.", False),
        ("", False),
        ("HOJA «Menu» — los departamentos, que son las pestañas del catálogo", True),
        ("   Nombre en la web   cómo quieres que se lea en la página", False),
        ("   Mostrar            SI lo enseña, NO lo esconde entero", False),
        ("   Orden              1 sale primero, 2 después, y así", False),
        ("", False),
        ("HOJA «Categorias» — los bloques dentro de cada departamento", True),
        ("   Cuantos productos enseñar   cuántos se listan antes del «y N más»", False),
        ("", False),
        ("HOJA «Productos» — el detalle, por clave", True),
        ("   Destacar   SI lo sube a los primeros de su categoría", False),
        ("", False),
        ("Cuando termines, guarda el archivo y mándamelo: se regenera la página.", False),
        ("Al volver a crear este archivo se conserva todo lo que ya decidiste;", False),
        ("sólo se añaden los renglones nuevos que hayan aparecido en el export.", False),
    ]
    for texto, fuerte in lineas:
        guia.append([texto])
        if fuerte:
            guia.cell(row=guia.max_row, column=1).font = Font(bold=True, size=12, color="8B3E1F")

    h1 = hoja(libro, "Menu", [
        ("Clave del departamento", 26), ("Productos", 11),
        ("Nombre en la web", 26), ("Mostrar", 11), ("Orden", 9),
    ], filas_menu, editables_desde=3)
    si_no(h1, "D", 2, h1.max_row)

    h2 = hoja(libro, "Categorias", [
        ("Clave del departamento", 24), ("Clave de la categoria", 30), ("Productos", 11),
        ("Nombre en la web", 30), ("Mostrar", 11), ("Cuantos productos enseñar", 26),
    ], filas_cat, editables_desde=4)
    si_no(h2, "E", 2, h2.max_row)

    h3 = hoja(libro, "Productos", [
        ("Clave", 12), ("Departamento", 20), ("Categoria", 26), ("Nombre en el punto de venta", 34),
        ("Nombre en la web", 34), ("Mostrar", 11), ("Destacar", 11),
    ], filas_prod, editables_desde=5)
    si_no(h3, "F", 2, h3.max_row)
    si_no(h3, "G", 2, h3.max_row)

    libro.save(CONTROL)
    print(f"{CONTROL.name}")
    print(f"  Menu:        {len(filas_menu)} departamentos")
    print(f"  Categorias:  {len(filas_cat)}")
    print(f"  Productos:   {len(filas_prod)}")
    if menu_previo:
        print("  (se conservaron las decisiones del archivo anterior)")


if __name__ == "__main__":
    main()
